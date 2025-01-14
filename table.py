import pandas as pd
import sqlite3
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def truncate_text(text, max_length):
    """Truncate a string and add ellipsis if it exceeds the max_length."""
    if isinstance(text, str) and len(text) > max_length:
        return text[:max_length - 3] + "..."
    return text

def parse_ticket_name(ticket_name):
    """
    Parse the multi-line ticket_name into separate components:
    Section Number, Row Letter, and View Type.
    Handles cases where words like 'tickets' are split across lines.
    """
    if not isinstance(ticket_name, str):
        return pd.Series([None, None, None])
    
    # Split the ticket_name by newline characters and strip whitespace
    lines = [line.strip() for line in ticket_name.split('\n') if line.strip()]
    
    section = None
    row = None
    view = None
    
    try:
        # Iterate through lines to find Section, Row, and View
        i = 0
        while i < len(lines):
            line = lines[i].lower()
            if line == 'section':
                if i + 1 < len(lines):
                    section = lines[i + 1]
                    i += 2
                    continue
            elif line == 'row':
                if i + 1 < len(lines):
                    row = lines[i + 1]
                    i += 2
                    continue
            elif 'view' in line:
                # Assume the entire line is the view type
                view = lines[i]
                i += 1
                continue
            else:
                # Handle split words like '2 ticket' + 's together'
                if 'ticket' in line and i + 1 < len(lines):
                    # Combine the current line with the next line to form 'tickets'
                    combined = line + lines[i + 1]
                    # Since we don't need 'ticket_arrangement', we skip processing it
                    i += 2
                    continue
            i += 1
        
        # Additional handling if 'View' is not explicitly labeled
        if not view and len(lines) >= 5:
            # Assume the last line is the view type
            view = lines[-1]
        
    except Exception as e:
        print(f"[DEBUG] Error parsing ticket_name '{ticket_name}': {e}")
    
    return pd.Series([section, row, view])

def export_to_excel_with_truncation(database_path, output_file, truncate_columns=None, max_length=30):
    """
    Export events and tickets data from SQLite database to an Excel file with styling.
    Truncate specified columns and add ellipses if the text exceeds max_length.
    The 'unique_id' and 'ticket_name' columns are removed from the tickets sheet before export.
    Includes the 'zone' column in the tickets sheet.
    Splits 'ticket_name' into separate columns: Section, Row, and View.
    Includes column titles (headers) in the Excel sheets.
    Adds a 'VIP Status' column with values 'yes' or 'no'.
    """
    # Connect to the SQLite database
    conn = sqlite3.connect(database_path)
    cursor = conn.cursor()
    
    # [MODIFICATION] Verify the presence of 'zone' column
    tickets_columns = [description[0] for description in cursor.execute("PRAGMA table_info(tickets)").fetchall()]
    if 'zone' not in tickets_columns:
        print("[WARNING] 'zone' column not found in 'tickets' table. Please ensure it exists before exporting.")
    
    try:
        # Load data from events and tickets tables
        events_df = pd.read_sql_query("SELECT * FROM events", conn)
        tickets_df = pd.read_sql_query("SELECT * FROM tickets", conn)
    
        # [MODIFICATION] Remove 'title' column from events if it exists
        if 'title' in events_df.columns:
            events_df = events_df.drop(columns=['title'])
            print("[DEBUG] 'title' column removed from 'events' table.")
    
        # [MODIFICATION] Remove 'title' column from tickets if it exists
        if 'title' in tickets_df.columns:
            tickets_df = tickets_df.drop(columns=['title'])
            print("[DEBUG] 'title' column removed from 'tickets' table.")
    
        # Remove 'unique_id' column from tickets if it exists
        if 'unique_id' in tickets_df.columns:
            tickets_df = tickets_df.drop(columns=['unique_id'])
            print("[DEBUG] 'unique_id' column removed from 'tickets' table.")
    
        # [MODIFICATION] Parse 'ticket_name' into separate columns: Section, Row, View
        tickets_df[['Section', 'Row', 'View']] = tickets_df['ticket_name'].apply(parse_ticket_name)
        print("[DEBUG] 'ticket_name' parsed into 'Section', 'Row', and 'View' columns.")
    
        # [MODIFICATION] Drop the original 'ticket_name' column
        tickets_df = tickets_df.drop(columns=['ticket_name'])
        print("[DEBUG] 'ticket_name' column removed from 'tickets' table.")
    
        # [MODIFICATION] Map 'is_vip' to 'yes'/'no'
        if 'is_vip' in tickets_df.columns:
            tickets_df['is_vip'] = tickets_df['is_vip'].map({1: 'yes', 0: 'no'})
            print("[DEBUG] 'is_vip' column mapped to 'yes'/'no'.")
        else:
            # If 'is_vip' column does not exist, add 'is_vip' as 'no' by default
            tickets_df['is_vip'] = 'no'
            print("[DEBUG] 'is_vip' column not found. Set all to 'no'.")
    
        # [MODIFICATION] Rearrange columns in 'tickets_df'
        # Define the desired order: Section, Row, View, zone, is_vip, followed by other columns, with event_link last
        desired_order = ['Section', 'Row', 'View', 'zone', 'is_vip']
        other_columns = [col for col in tickets_df.columns if col not in desired_order]
        # Ensure 'event_link' is last
        if 'event_link' in other_columns:
            other_columns.remove('event_link')
            other_columns.append('event_link')
        tickets_df = tickets_df[desired_order + other_columns]
        print("[DEBUG] Columns in 'tickets' table rearranged.")
    
        # [MODIFICATION] Verify that 'ticket_name' has been removed
        if 'ticket_name' in tickets_df.columns:
            print("[ERROR] 'ticket_name' column still exists in 'tickets_df'. Please check the drop operation.")
        else:
            print("[DEBUG] 'ticket_name' column successfully removed from 'tickets_df'.")
    
        # Apply truncation if needed
        if truncate_columns:
            for col in truncate_columns:
                if col in events_df.columns:
                    events_df[col] = events_df[col].apply(lambda x: truncate_text(x, max_length))
                    print(f"[DEBUG] Truncated column '{col}' in 'events' table.")
                if col in tickets_df.columns:
                    tickets_df[col] = tickets_df[col].apply(lambda x: truncate_text(x, max_length))
                    print(f"[DEBUG] Truncated column '{col}' in 'tickets' table.")
    
        # [MODIFICATION] Print columns of tickets_df for verification
        print(f"[DEBUG] Final columns in 'tickets_df': {tickets_df.columns.tolist()}")
    
        # Create a new Excel workbook
        wb = Workbook()
    
        # Add the 'Events' sheet with headers
        ws_events = wb.active
        ws_events.title = "Events"
        for row in dataframe_to_rows(events_df, index=False, header=True):  # [MODIFICATION] header=True
            ws_events.append(row)
        print("[DEBUG] 'Events' sheet populated with headers.")
    
        # Add the 'Tickets' sheet with headers
        ws_tickets = wb.create_sheet(title="Tickets")
        for row in dataframe_to_rows(tickets_df, index=False, header=True):  # [MODIFICATION] header=True
            ws_tickets.append(row)
        print("[DEBUG] 'Tickets' sheet populated with headers.")
    
        # Apply styling to both sheets
        for sheet in [ws_events, ws_tickets]:
            # Bold headers, center alignment, and wrap text
            for cell in sheet[1]:  # First row contains headers
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
            # Alternating row colors and text wrapping
            fill_gray = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
            for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                                                    min_col=1, max_col=sheet.max_column), start=2):
                fill = fill_gray if i % 2 == 0 else fill_white
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True)
    
            # Auto-adjust column widths
            for col in sheet.columns:
                col_letter = get_column_letter(col[0].column)
                # Calculate the maximum length of the column's data
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                # Set the column width with a buffer
                sheet.column_dimensions[col_letter].width = max_len + 2
        print("[DEBUG] Styling applied to both sheets.")
    
        # Save the workbook
        wb.save(output_file)
        print(f"Data successfully exported to {output_file} with truncation and no 'unique_id' or 'ticket_name' columns.")
    except Exception as e:
        print(f"Error exporting data: {e}")
    finally:
        conn.close()
        print("[DEBUG] Database connection closed.")

# Example usage
if __name__ == "__main__":
    database_path = "events.db"
    output_file = "styled_event_ticket_data_truncated.xlsx"
    truncate_columns = ["zone"]  # [MODIFICATION] Updated columns to truncate
    max_length = 40  # Maximum length for truncated columns

    export_to_excel_with_truncation(database_path, output_file, truncate_columns, max_length)
